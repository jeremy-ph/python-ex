import openpyxl
# pip install openpyxl  (라이브러리 설치)


wb = openpyxl.load_workbook('./sample.xlsx')
sheet = wb['Sheet1']

# # ex 1 
# print(sheet.max_column, sheet.max_row)
# print(sheet.cell(row=1, column=1).value)
# print(sheet.cell(row=2, column=1).value)

# # ex 2
# for row in sheet.iter_rows(min_row=2):
#     for cell in row:
#         print(cell.value)
#     print("-" * 10)

# ex 3
cells = sheet['A2':'C4']
for row in cells:
    for cell in row:
        print(cell.value)
    print("-" * 10)

wb.close()

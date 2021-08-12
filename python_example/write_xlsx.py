import openpyxl
# 엑셀 파일 만들기

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = '회원정보'

# 표 헤더 컬럼 저장
header_titles = ['이름', '전화번호']
for idx, title in enumerate(header_titles):
    sheet.cell(row=1, column=idx+1, value=title)

# 표 내용 저장
numbers = [('jeremy', '010-000-2222'), ('hong', '010-222-3333')]
row_num = 2 # 1번째 row는 타이틀 위치
for r, number in enumerate(numbers):    # 회원정보 목록 탐색
    for c, v in enumerate(number):
        sheet.cell(row=row_num+r, column=c+1, value=v)

wb.save('./members.xlsx')
wb.close